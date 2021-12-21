from openpyxl import Workbook, load_workbook

wb = load_workbook('data.xlsx')

ws = wb.active

print(ws['A2'].value)
